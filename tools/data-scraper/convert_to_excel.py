#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convert TikTok JSON data to CSV and Excel format using openpyxl
"""
import json
import csv
from datetime import datetime

# Load JSON data
with open('tiktok_apify_data/videos_full.json', 'r', encoding='utf-8') as f:
    videos = json.load(f)

# Extract key fields
data = []
for video in videos:
    row = [
        video.get('id', ''),
        video.get('createTimeISO', ''),
        (video.get('text', '')[:100] if video.get('text') else ''),
        video.get('diggCount', 0) or video.get('hearts', 0),
        video.get('commentCount', 0) or video.get('comments', 0),
        video.get('shareCount', 0) or video.get('shares', 0),
        video.get('playCount', 0),
        video.get('webVideoUrl', ''),
        video.get('musicMeta', {}).get('musicName', '') if video.get('musicMeta') else '',
        video.get('textLanguage', ''),
        'Yes' if video.get('isAd') else 'No',
    ]
    data.append(row)

# Sort by publish time (newest first)
data.sort(key=lambda x: x[1], reverse=True)

# Save to CSV first
csv_path = 'tiktok_apify_data/BYD_AUTO_BRASIL_Videos.csv'
headers = ['视频ID', '发布时间', '视频标题', '点赞数', '评论数', '分享数', '播放次数',
           '视频URL', '音乐', '语言', '是否广告']

with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    writer.writerows(data)

print(f"✅ CSV 文件已生成: {csv_path}")

# Now convert CSV to Excel using openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '视频数据'

    # Write headers
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data and adjust column widths
    for row in data:
        ws.append(row)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 8

    # Save Excel file
    excel_path = 'tiktok_apify_data/BYD_AUTO_BRASIL_Videos.xlsx'
    wb.save(excel_path)
    print(f"✅ Excel 文件已生成: {excel_path}")

except ImportError:
    print("⚠️ openpyxl 未安装，但 CSV 文件已生成")

